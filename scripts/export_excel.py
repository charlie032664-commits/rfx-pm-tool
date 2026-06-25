# -*- coding: utf-8 -*-
"""
Export compliance matrix Excel (3 sheets) from requirements_clean.json
(output of postprocess_requirements.py). Falls back to requirements_enriched.json (legacy).

Usage:
  python export_excel.py --case sample_case
  python export_excel.py --in runs/sample_case/requirements_clean.json --out runs/sample_case/compliance_matrix.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


# ── v1.4 Excel UX Phase 1: PM internal review column order (rename + reorder).
#    See docs/v1.4_excel_output_ux_spec.md §3. PM-facing requirement + response
#    columns come first; AI/normalization audit fields move to a right-hand block.
#    This is display order/labels only — underlying values are unchanged.
HEADERS = [
    "Req ID",
    "Priority",
    "Category",
    "Responsible Team",
    "Stakeholder",
    "Compliance Status",
    # ── PM-facing requirement + response flow ──
    "PM Reviewed Requirement",          # was "Requirement (Final)"
    "Customer Requirement (Original)",  # was "Requirement (Original)"
    "Our Response",
    "Gap / Notes",
    "Evidence",
    "Risk Tags",
    "Source",
    # ── AI internal audit block (right side; filled by normalize_requirements_llm.py) ──
    "AI Parsed Requirement",            # was "Requirement (Normalized)"
    "AI Rewrite Status",                # was "Rewrite Reason"
    "AI Confidence",                    # was "Rewrite Confidence"
    "PM AI Review Status",              # was "Rewrite Review"
]

HIDDEN_COLS: set = set()


# ── Big category buckets ─────────────────────────────────────────────────────
# Maps small category (case-insensitive) → big-category sheet name.
# Multi-value categories (e.g. "Compliance, Security") use only the FIRST value
# to avoid double counting. Blank or unmapped categories fall through to "Others".

BIG_CATEGORY_MAP: Dict[str, str] = {
    # ── Hardware ────────────────────────────────────────────────────────────
    "hardware":      "Hardware",
    "platform":      "Hardware",
    "cpu":           "Hardware",
    "memory":        "Hardware",
    "pcie":          "Hardware",
    "storage":       "Hardware",
    "storage/hw":    "Hardware",
    "network":       "Hardware",
    "wireless":      "Hardware",
    "lan":           "Hardware",
    "display":       "Hardware",
    "interface":     "Hardware",
    "i/o":           "Hardware",
    "io":            "Hardware",
    "power":         "Hardware",
    "tpm":           "Hardware",
    "performance":   "Hardware",
    # ── Software ────────────────────────────────────────────────────────────
    "software":      "Software",
    "bios":          "Software",
    "bmc":           "Software",
    "firmware":      "Software",
    "driver":        "Software",
    "os":            "Software",
    "management":    "Software",
    "storage/sw":    "Software",
    # ── Mechanical ──────────────────────────────────────────────────────────
    "mechanical":     "Mechanical",
    "serviceability": "Mechanical",
    "chassis":        "Mechanical",
    "bezel":          "Mechanical",
    "dimension":      "Mechanical",
    "label":          "Mechanical",
    "packaging":      "Mechanical",
    # ── Regulatory ──────────────────────────────────────────────────────────
    "regulatory":    "Regulatory",
    "emc":           "Regulatory",
    "safety":        "Regulatory",
    "ce":            "Regulatory",
    "fcc":           "Regulatory",
    "rohs":          "Regulatory",
    "certification": "Regulatory",
    "compliance":    "Regulatory",
    "security":      "Regulatory",
    "legal":         "Regulatory",
    # ── Environmental ───────────────────────────────────────────────────────
    "environmental": "Environmental",
    "temperature":   "Environmental",
    "thermal":       "Environmental",
    "humidity":      "Environmental",
    "shock":         "Environmental",
    "vibration":     "Environmental",
    "reliability":   "Environmental",
    # ── Others (explicit; blank/unmapped also default to Others) ────────────
    "general":       "Others",
    "commercial":    "Others",
    "documentation": "Others",
    "unknown":       "Others",
}

BIG_CATEGORY_ORDER = [
    "Hardware", "Software", "Mechanical", "Regulatory", "Environmental", "Others",
]


def _first_category(category) -> str:
    """Return the first category from a list or comma-joined string."""
    if isinstance(category, list):
        return str(category[0]).strip() if category else ""
    s = str(category or "").strip()
    if "," in s:
        return s.split(",", 1)[0].strip()
    return s


def _to_big_category(category) -> str:
    """Map first-of small category to its big-category bucket. Unknown/blank → Others."""
    first = _first_category(category)
    return BIG_CATEGORY_MAP.get(first.lower(), "Others")


def _has_redflag(r: Dict[str, Any]) -> bool:
    msgs = r.get("risk_tags") or r.get("redflag_tags") or r.get("redflag_messages") or []
    return bool(msgs and str(msgs) not in ("[]", ""))


def bucket_by_big_category(reqs: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """Group requirements into the 6 big-category buckets.

    All buckets are initialized (may be empty) so that empty-sheet creation
    downstream is deterministic. Preserves caller-provided ordering within
    each bucket (so a pre-sorted main_reqs stays sorted inside each sheet).
    """
    buckets: Dict[str, List[Dict[str, Any]]] = {bc: [] for bc in BIG_CATEGORY_ORDER}
    for r in reqs:
        big = _to_big_category(r.get("category", ""))
        buckets[big].append(r)
    return buckets


def write_summary_sheet(ws, buckets: Dict[str, List[Dict[str, Any]]]) -> None:
    """Write a pivot of counts per (Group, small Category).

    Columns: Group | Category | Total | MUST | SHOULD | MAY | INFO
             | NEED_REVIEW | COMPLIANT | PARTIAL | NON_COMPLIANT | Owner(s)

    Rows are one per (Group, small Category) actually present in the data.
    Multi-value categories use the FIRST value (same rule as bucket_by_big_category).
    A final "Total" grand-row sums all numeric columns.
    """
    headers = [
        "Group", "Category", "Total",
        "MUST", "SHOULD", "MAY", "INFO",
        "NEED_REVIEW", "COMPLIANT", "PARTIAL", "NON_COMPLIANT",
        "Owner(s)",
    ]
    ws.append(headers)

    grand = {"total": 0, "must": 0, "should": 0, "may": 0, "info": 0,
             "nr": 0, "comp": 0, "part": 0, "non": 0}

    for big_cat in BIG_CATEGORY_ORDER:
        items = buckets.get(big_cat, [])
        if not items:
            continue
        # Sub-group by small category (first-of), stable alphabetical order
        by_small: Dict[str, List[Dict[str, Any]]] = {}
        for r in items:
            small = _first_category(r.get("category", "")) or "(blank)"
            by_small.setdefault(small, []).append(r)
        for small in sorted(by_small.keys(), key=str.lower):
            sub = by_small[small]
            total  = len(sub)
            must   = sum(1 for r in sub if str(r.get("must_level", "")).upper() == "MUST")
            should = sum(1 for r in sub if str(r.get("must_level", "")).upper() == "SHOULD")
            may    = sum(1 for r in sub if str(r.get("must_level", "")).upper() == "MAY")
            info   = sum(1 for r in sub if str(r.get("must_level", "")).upper() == "INFO")
            # NOTE: status COMPLIANT/PARTIAL/NON-COMPLIANT come from responses.json,
            # already merged into r["status"] above. The status string uses a HYPHEN
            # ("NON-COMPLIANT"); the column header keeps underscore as requested.
            status_u = lambda r: str(r.get("status", "")).upper()
            nr     = sum(1 for r in sub if status_u(r) == "NEED_REVIEW")
            comp   = sum(1 for r in sub if status_u(r) == "COMPLIANT")
            part   = sum(1 for r in sub if status_u(r) == "PARTIAL")
            non    = sum(1 for r in sub if status_u(r) == "NON-COMPLIANT")
            owners = sorted({
                str(r.get("owner", "")).strip()
                for r in sub
                if str(r.get("owner", "")).strip()
            }, key=str.lower)
            ws.append([big_cat, small, total, must, should, may, info,
                       nr, comp, part, non, ", ".join(owners)])
            grand["total"]  += total
            grand["must"]   += must
            grand["should"] += should
            grand["may"]    += may
            grand["info"]   += info
            grand["nr"]     += nr
            grand["comp"]   += comp
            grand["part"]   += part
            grand["non"]    += non

    # Grand-total row (Owner(s) column intentionally blank)
    ws.append([
        "Total", "",
        grand["total"], grand["must"], grand["should"], grand["may"], grand["info"],
        grand["nr"], grand["comp"], grand["part"], grand["non"],
        "",
    ])

    # ── Styling ────────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    numeric_cols = {3, 4, 5, 6, 7, 8, 9, 10, 11}
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col in numeric_cols:
                cell.alignment = Alignment(vertical="center", horizontal="right")
            else:
                cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    # Bold + tinted grand-total row (last row)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True, name="Arial")
        cell.fill = PatternFill("solid", fgColor="E3F2FD")

    widths = {
        1: 14,   # Group
        2: 18,   # Category
        3: 8,    # Total
        4: 8,    # MUST
        5: 9,    # SHOULD
        6: 7,    # MAY
        7: 7,    # INFO
        8: 14,   # NEED_REVIEW
        9: 12,   # COMPLIANT
        10: 10,  # PARTIAL
        11: 16,  # NON_COMPLIANT
        12: 32,  # Owner(s)
    }
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "C2"


def read_json(p: Path) -> Dict[str, Any]:
    return json.loads(p.read_text(encoding="utf-8"))


def _category_str(category) -> str:
    if isinstance(category, list):
        return ", ".join(map(str, category))
    return str(category) if category is not None else ""


def _source_ref(r: Dict[str, Any]) -> str:
    import re
    s = r.get("source")

    def _clean_name(f):
        f = re.sub(r'\.(docx|doc|xlsx|xls|pdf|txt|md)$', '', str(f), flags=re.IGNORECASE)
        f = re.sub(r'^AUTO[-_]?', '', f, flags=re.IGNORECASE)
        f = re.sub(r'^(UNKNOWN|UnknownFile)$', '', f, flags=re.IGNORECASE)
        return f.strip()[:40]

    # dict 格式（requirements_enriched.json）
    if isinstance(s, dict):
        f = _clean_name(s.get("file") or "")
        if not f:
            return ""
        sheet = s.get("sheet") or ""
        row = s.get("row")
        table = s.get("table")
        table_row = s.get("table_row")
        page = s.get("page")
        chunk = s.get("chunk")
        if sheet and row is not None:
            return f"{f} \u2014 Sheet: {sheet}, \u7b2c {row} \u884c"
        if sheet:
            return f"{f} \u2014 Sheet: {sheet}"
        if table is not None and table_row is not None:
            return f"{f} \u2014 Table {table}, Row {table_row}"
        if page is not None:
            return f"{f} \u2014 \u7b2c {page} \u9801"
        if chunk is not None:
            return f"{f} \u2014 \u7b2c {chunk} \u6bb5"
        return f

    # 字串格式（requirements_clean.json）
    # 格式可能是：
    #   "TSR 2U 2S... — Table 2, Row 4"  (含位置資訊，直接使用)
    #   "TSR 2U 2S...#chunk5"             (舊格式，需解析)
    #   "TSR 2U 2S..."                    (只有檔名)
    if isinstance(s, str) and s:
        # 若已含有 — 分隔符，表示是格式化好的字串，直接使用
        if ' \u2014 ' in s:
            return s[:80]
        # 舊的 #chunk 格式
        if '#' in s:
            file_part, _, chunk_part = s.partition('#')
            file_part = _clean_name(file_part)
            if not file_part:
                return ""
            chunk_num = re.search(r'\d+', chunk_part)
            n = chunk_num.group() if chunk_num else chunk_part
            return f"{file_part} \u2014 \u7b2c {n} \u6bb5"
        # 只有檔名
        return _clean_name(s)

    return ""


def _load_reqs(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Accept reviewed (v1.4), clean (items), and enriched (requirements) schemas.

    Priority of detection:
      1. v1.4 PM-reviewed schema  — meta.schema_version == "v1.4_review_mvp"
      2. requirements_clean.json  — top-level "items"
      3. requirements_enriched.json (legacy) — top-level "requirements"
    """
    # ── v1.4 PM-reviewed schema (requirements_reviewed.json) ──────────────────
    # Detect by meta.schema_version, then map the reviewed row shape onto the
    # item shape the rest of the export flow expects (req_id / requirement /
    # category / must_level / owner). Rows flagged deleted are skipped here, and
    # pm_comment is intentionally NOT carried into the customer-facing matrix.
    if data.get("meta", {}).get("schema_version") == "v1.4_review_mvp":
        out = []
        for req in (data.get("requirements", []) or []):
            if req.get("deleted"):
                continue
            rid = req.get("id", "")
            r = {
                "req_id":       rid,
                "orig_req_id":  rid,
                "requirement":  req.get("requirement", ""),
                "category":     req.get("category", ""),
                "must_level":   req.get("must_level", ""),
                "owner":        req.get("owner", ""),
                "risk_tags":    [],
            }
            out.append(r)
        return out
    if "items" in data:
        out = []
        for item in (data["items"] or []):
            r = dict(item)
            # Normalize: ensure risk_tags is the canonical field
            if "risk_tags" not in r:
                r["risk_tags"] = r.get("redflag_tags") or r.get("redflag_messages") or []
            out.append(r)
        return out
    # Enriched format: redflag_messages → risk_tags
    out = []
    for req in (data.get("requirements", []) or []):
        r = dict(req)
        if "risk_tags" not in r:
            r["risk_tags"] = r.get("redflag_tags") or r.get("redflag_messages") or []
        out.append(r)
    return out


def split_sheets(reqs: List[Dict[str, Any]]) -> Tuple[List, List, List, List, List]:
    """Route items into 5 buckets — main / glossary / notes / skipped / pm_excluded.

    Routing precedence (first match wins):
      0. PM Excluded — exclude_from_matrix is True (Phase 4.6E.2; PM-explicit
                       decision wins over every type-based routing below)
      1. Glossary   — type == "glossary"  OR  risk_tags contains "GLOSSARY"
      2. Notes      — type == "note"
      3. Skipped    — type == "junk"  OR  status == "AUTO_SKIP"
                      (catches obvious-noise rows: dates, names, ticks, etc.,
                       and any AUTO_SKIP rows that weren't glossary/note —
                       e.g., short colon-prefixed values the enricher
                       mis-tagged.)
      4. Main       — everything else (real requirements)
    """
    main, glossary, notes, skipped, pm_excluded = [], [], [], [], []
    for r in reqs:
        # Phase 4.6E.2 — PM explicit exclude takes highest priority
        if r.get("exclude_from_matrix") is True:
            pm_excluded.append(r)
            continue

        req_type = str(r.get("type") or "").lower()
        status   = str(r.get("status") or "").upper()
        raw_tags = r.get("risk_tags") or []
        if not isinstance(raw_tags, list):
            raw_tags = [raw_tags]
        tag_set = {str(t).upper() for t in raw_tags}

        if req_type == "glossary" or "GLOSSARY" in tag_set:
            glossary.append(r)
            continue
        if req_type == "note":
            notes.append(r)
            continue
        if req_type == "junk" or status == "AUTO_SKIP":
            skipped.append(r)
            continue
        main.append(r)

    return main, glossary, notes, skipped, pm_excluded


def apply_sheet_style(ws) -> None:
    header_font  = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    top  = Alignment(vertical="top")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.freeze_panes = "A2"

    wrap_cols = {
        "PM Reviewed Requirement",
        "Customer Requirement (Original)",
        "AI Parsed Requirement",
        "Stakeholder",
        "Risk Tags",
        "Our Response",
        "Gap / Notes",
        "Evidence",
    }

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            h = ws.cell(row=1, column=col).value
            cell.alignment = wrap if h in wrap_cols else top

    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    col_width = {h: 18 for h in HEADERS}
    col_width.update({
        "Req ID": 15,
        "Priority": 12,
        "Category": 15,
        "Responsible Team": 14,
        "Stakeholder": 22,
        "Compliance Status": 18,
        "PM Reviewed Requirement": 60,
        "Customer Requirement (Original)": 60,
        "AI Parsed Requirement": 60,
        "AI Rewrite Status": 24,
        "AI Confidence": 12,
        "PM AI Review Status": 10,
        "Risk Tags": 22,
        "Our Response": 50,
        "Gap / Notes": 35,
        "Evidence": 35,
        "Source": 40,
    })
    for idx, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col_width.get(h, 18)
        if h in HIDDEN_COLS:
            ws.column_dimensions[get_column_letter(idx)].hidden = True

    # Conditional formatting
    if ws.max_row >= 2:
        last_col = get_column_letter(ws.max_column)
        data_range = f"A2:{last_col}{ws.max_row}"

        # ── Phase 4.6 normalization highlighting ───────────────────────────
        # Insert FIRST + stopIfTrue=True so the column-specific Normalized /
        # Review colour wins over the row-wide Status colour added below.
        if "AI Parsed Requirement" in HEADERS and "PM AI Review Status" in HEADERS:
            nc  = HEADERS.index("AI Parsed Requirement") + 1
            ncl = get_column_letter(nc)
            rc  = HEADERS.index("PM AI Review Status") + 1
            rcl = get_column_letter(rc)
            # Normalized cell → light yellow when row is flagged REVIEW
            ws.conditional_formatting.add(
                f"{ncl}2:{ncl}{ws.max_row}",
                FormulaRule(formula=[f'${rcl}2="REVIEW"'],
                            fill=PatternFill("solid", fgColor="FFE699"),
                            stopIfTrue=True)
            )
            # Normalized cell → light green when normalized has content and not REVIEW
            ws.conditional_formatting.add(
                f"{ncl}2:{ncl}{ws.max_row}",
                FormulaRule(formula=[f'AND(${ncl}2<>"",${rcl}2<>"REVIEW")'],
                            fill=PatternFill("solid", fgColor="DDF0E0"),
                            stopIfTrue=True)
            )
            # Review cell → bold yellow when "REVIEW"
            ws.conditional_formatting.add(
                f"{rcl}2:{rcl}{ws.max_row}",
                FormulaRule(formula=[f'${rcl}2="REVIEW"'],
                            fill=PatternFill("solid", fgColor="FFD966"),
                            stopIfTrue=True)
            )

        # ── Status row colouring (lower priority; added AFTER Normalized) ──
        sc = HEADERS.index("Compliance Status") + 1
        sl = get_column_letter(sc)
        _status_colors = [
            ("NEED_REVIEW",        "FFF2CC"),  # warm yellow
            ("NEW",                "E3F2FD"),  # light blue
            ("INTERNAL_ALIGN",     "E8EAF6"),  # light indigo
            ("ASK_CUSTOMER",       "FFF3E0"),  # light orange
            ("READY_FOR_RESPONSE", "EAF3DE"),  # light green
            ("CLOSED",             "F5F5F5"),  # light grey
            ("AUTO_SKIP",          "E7E6E6"),  # grey
        ]
        for _sv, _clr in _status_colors:
            ws.conditional_formatting.add(data_range, FormulaRule(
                formula=[f'${sl}2="{_sv}"'],
                fill=PatternFill("solid", fgColor=_clr)
            ))
        mc = HEADERS.index("Priority") + 1
        ml = get_column_letter(mc)
        ws.conditional_formatting.add(
            f"{ml}2:{ml}{ws.max_row}",
            FormulaRule(formula=[f'${ml}2="MUST"'], fill=PatternFill("solid", fgColor="F8CBAD"))
        )


def write_sheet(ws, reqs: List[Dict[str, Any]]) -> None:
    ws.append(HEADERS)

    for r in reqs:
        req_id      = r.get("req_id", "")
        must_level  = r.get("must_level", "")
        category    = _category_str(r.get("category", []))
        owner       = r.get("owner", "")
        # Stakeholder — list of additional involved teams, comma-joined
        _sh = r.get("stakeholder") or []
        if isinstance(_sh, list):
            stakeholder = ", ".join(str(s).strip() for s in _sh if str(s).strip())
        else:
            stakeholder = str(_sh).strip()
        status      = r.get("status", "")
        req_text    = r.get("requirement", "")
        source      = _source_ref(r)

        risk_tags = r.get("risk_tags") or r.get("redflag_tags") or r.get("redflag_messages") or []
        risk_str = ", ".join(risk_tags) if isinstance(risk_tags, list) else str(risk_tags)

        our_response = r.get("vendor_comment") or r.get("our_response") or ""
        gap_notes    = r.get("gap") or ""
        evidence     = r.get("evidence_needed") or r.get("evidence") or ""

        # ── Phase 4.6 normalization fields ──
        normalized     = (r.get("normalized_requirement") or "").strip()
        rewrite_reason = (r.get("rewrite_reason") or "").strip()
        rewrite_conf   = r.get("rewrite_confidence", 0.0)
        needs_review   = bool(r.get("needs_rewrite_review", False))
        # Confidence shows "0.00" only when normalize was attempted; blank otherwise
        if rewrite_reason:
            try:
                conf_str = f"{float(rewrite_conf):.2f}"
            except (TypeError, ValueError):
                conf_str = "0.00"
        else:
            conf_str = ""
        review_str = "REVIEW" if needs_review else ""

        # ── Phase 4.6E.2: Requirement (Final) fallback chain ──
        #     PM edit  →  LLM normalized  →  Original
        _pm_final = (r.get("final_requirement") or "").strip()
        if _pm_final:
            final_req = _pm_final
        elif normalized:
            final_req = normalized
        else:
            final_req = req_text

        # Order MUST match HEADERS (v1.4 Excel UX Phase 1). Same values as before;
        # only the column positions/labels changed. AI fields moved to the right.
        ws.append([
            req_id, must_level, category, owner, stakeholder, status,
            final_req,        # PM Reviewed Requirement
            req_text,         # Customer Requirement (Original)
            our_response, gap_notes, evidence,
            risk_str,         # Risk Tags
            source,
            normalized,       # AI Parsed Requirement
            rewrite_reason,   # AI Rewrite Status
            conf_str,         # AI Confidence
            review_str,       # PM AI Review Status
        ])

    apply_sheet_style(ws)

    # Colour-code Req ID: RFQ- green, AI- blue
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        val = str(cell.value or "")
        if val.startswith("RFQ-"):
            cell.fill = PatternFill("solid", fgColor="EAF3DE")
            cell.font = Font(name="Arial", size=10, color="27500A")
        elif val.startswith("AI-"):
            cell.fill = PatternFill("solid", fgColor="E6F1FB")
            cell.font = Font(name="Arial", size=10, color="0C447C")


def _merge_responses(reqs: List[Dict[str, Any]], responses: Dict[str, Any]) -> None:
    """Apply PM responses onto items in-place (Phase 4.6E.2).

    Runs BEFORE split_sheets so that `exclude_from_matrix=True` can drive
    routing into the dedicated 'Excluded' sheet. Items without a matching
    response entry are untouched.

    For excluded rows, the Gap / Notes column is prefixed with
    "[EXCLUDED: <reason>]" (or "[EXCLUDED]" if no reason was given) so the
    Excluded sheet can show the reason without a separate column.
    """
    for r in reqs:
        resp = responses.get(r.get("req_id", ""), {})
        if not resp:
            continue
        if resp.get("status"):           r["status"]            = resp["status"]
        if resp.get("vendor_comment"):   r["vendor_comment"]    = resp["vendor_comment"]
        if resp.get("gap"):              r["gap"]               = resp["gap"]
        if resp.get("ai_draft"):         r["ai_draft"]          = resp["ai_draft"]
        # Phase 4.6E.2 — three new fields
        if resp.get("final_requirement"):
            r["final_requirement"] = resp["final_requirement"]
        r["exclude_from_matrix"] = bool(resp.get("exclude_from_matrix", False))
        r["exclude_reason"]      = str(resp.get("exclude_reason", "") or "")

    # Prefix [EXCLUDED] tag to Gap / Notes for excluded rows. Done in a second
    # pass so the source `r["gap"]` is the freshly-merged value, not stale.
    for r in reqs:
        if r.get("exclude_from_matrix") is True:
            reason = (r.get("exclude_reason") or "").strip()
            tag    = f"[EXCLUDED: {reason}]" if reason else "[EXCLUDED]"
            gap    = (r.get("gap") or "").strip()
            r["gap"] = f"{tag} {gap}" if gap else tag


def export_excel(data: Dict[str, Any], out_xlsx: Path, responses_path: Optional[Path] = None) -> None:
    # Load responses.json (owner-filled data)
    responses: Dict[str, Any] = {}
    if responses_path and Path(responses_path).exists():
        try:
            responses = json.loads(Path(responses_path).read_text(encoding="utf-8"))
            print(f"[INFO] Loaded {len(responses)} responses from {responses_path}")
        except Exception as e:
            print(f"[WARN] Failed to load responses: {e}")

    wb = Workbook()
    wb.remove(wb.active)

    reqs = _load_reqs(data)
    # Phase 4.6E.2: merge responses BEFORE split so exclude_from_matrix routes
    # rows into the dedicated Excluded sheet (replaces the legacy post-split
    # merge that only touched main_reqs).
    _merge_responses(reqs, responses)
    main_reqs, glossary_reqs, notes_reqs, skipped_reqs, pm_excluded_reqs = split_sheets(reqs)

    # ── PM 排序：NEED_REVIEW 先 > 有 Redflag 先 > MUST 先 > Category ──────────
    _status_order = {"NEED_REVIEW": 0, "INTERNAL_ALIGN": 1, "ASK_CUSTOMER": 2,
                     "NEW": 3, "READY_FOR_RESPONSE": 4, "CLOSED": 5, "AUTO_SKIP": 6}
    _must_order   = {"MUST": 0, "SHOULD": 1, "MAY": 2, "INFO": 3}
    _cat_order    = {
        "Compliance": 0, "Reliability": 1, "Security": 2,
        "Platform": 3, "BMC": 4, "BIOS": 5, "Storage": 6,
        "PCIe": 7, "Power": 8, "Thermal": 9, "Mechanical": 10, "General": 99,
    }
    def _sort_key(r):
        msgs = r.get("risk_tags") or r.get("redflag_tags") or r.get("redflag_messages") or []
        has_rf = 0 if (msgs and str(msgs) not in ("[]", "")) else 1
        return (
            _status_order.get(str(r.get("status", "")), 9),
            has_rf,
            _must_order.get(str(r.get("must_level", "")), 9),
            _cat_order.get(_category_str(r.get("category", [])).split(",")[0].strip(), 50),
        )
    main_reqs = sorted(main_reqs, key=_sort_key)

    # Sheet order: Compliance Matrix (總表, sheet 1) → By_Category_Summary (sheet 2)
    # → 6 big-category sheets numbered 3..8 → Glossary → Notes.
    # The display name is numbered ("3. Hardware") for customer-facing readability,
    # but the internal Group label inside By_Category_Summary stays "Hardware".
    write_sheet(wb.create_sheet("Compliance Matrix"), main_reqs)

    # Bucket the already-sorted main_reqs so each big-category sheet preserves
    # the same sort order (NEED_REVIEW > redflag > MUST > category).
    buckets = bucket_by_big_category(main_reqs)
    write_summary_sheet(wb.create_sheet("By_Category_Summary"), buckets)
    for idx, big_cat in enumerate(BIG_CATEGORY_ORDER, start=3):
        sheet_name = f"{idx}. {big_cat}"
        write_sheet(wb.create_sheet(sheet_name), buckets[big_cat])

    write_sheet(wb.create_sheet("Glossary"), glossary_reqs)
    write_sheet(wb.create_sheet("Notes"),    notes_reqs)
    write_sheet(wb.create_sheet("Skipped"),  skipped_reqs)
    # Phase 4.6E.2 — PM-excluded sheet (placed after Skipped)
    write_sheet(wb.create_sheet("Excluded"), pm_excluded_reqs)

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def find_latest_clean(runs_dir: Path) -> Optional[Path]:
    for pattern in ("**/requirements_clean.json", "**/requirements_enriched.json"):
        candidates = list(runs_dir.glob(pattern))
        if candidates:
            candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            return candidates[0]
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in",        dest="in_path",       default=None)
    ap.add_argument("--out",       dest="out_path",      default=None)
    ap.add_argument("--case",      dest="case_id",       default=None)
    ap.add_argument("--responses", dest="responses_path", default=None)
    args = ap.parse_args()

    base_dir = Path(__file__).resolve().parent
    runs_dir = base_dir / "runs"

    in_path: Optional[Path] = None
    if args.in_path:
        in_path = Path(args.in_path)
        if not in_path.is_absolute():
            in_path = (base_dir / in_path).resolve()
    elif args.case_id:
        clean    = runs_dir / args.case_id / "requirements_clean.json"
        enriched = runs_dir / args.case_id / "requirements_enriched.json"
        in_path  = clean if clean.exists() else enriched
    else:
        in_path = find_latest_clean(runs_dir)

    if not in_path or not in_path.exists():
        raise FileNotFoundError(
            f"requirements_clean.json not found.\n"
            f"Run postprocess_requirements.py first, or specify --in <path>"
        )

    out_path = Path(args.out_path) if args.out_path else in_path.parent / "compliance_matrix.xlsx"
    if not out_path.is_absolute():
        out_path = (base_dir / out_path).resolve()

    data = read_json(in_path)
    responses_path = Path(args.responses_path) if args.responses_path else None
    export_excel(data, out_path, responses_path=responses_path)

    print(f"[OK] Input : {in_path}")
    print(f"[OK] Output: {out_path}")
    print("[OK] Sheets: Compliance Matrix, By_Category_Summary, "
          "3. Hardware, 4. Software, 5. Mechanical, "
          "6. Regulatory, 7. Environmental, 8. Others, "
          "Glossary, Notes, Skipped, Excluded")


if __name__ == "__main__":
    main()
