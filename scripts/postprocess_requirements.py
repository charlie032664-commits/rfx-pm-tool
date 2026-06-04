# -*- coding: utf-8 -*-
"""
Post-process requirements(.json or _enriched.json) -> clean.json + review.xlsx (3 sheets)

Excel Columns:
1  Req ID
2  重要程度 (Must Level)
3  Category
4  Owner
5  Stakeholder
6  Status
7  Requirement (Original)
8  Risk Flags / 風險標記
9  Evidence Notes
10 Next Action
11 Source  (hidden, format: Overview#chunk1)

Changes v3:
- Req ID: AI-001 (sequential) / RFQ-BMC-001 (from original table)
- Removed: Source Type, Source File, Chunk, Source Short, Excerpt
- Added: Source (single hidden column, format file#chunkN)
- Req ID colour-coded in Excel: AI- blue, RFQ- green
"""

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

import pandas as pd


PLACEHOLDER_VALUES = {"<file>", "<chunk>", "", None}

# RFQ table row ID patterns — these come from the original customer table
_RFQ_TABLE_ID_PATTERNS = [
    re.compile(r"^(HOST|BMC|BIOS|BIOS-BMC|PCIe|STORAGE|PWR)-\d+$", re.IGNORECASE),
    re.compile(r"^ROW_ID=\d+$"),
]

# Global AI sequence counter — used for *all* system-generated ids.
# Phase 4 ID policy (revised):
#   - RFQ-* : customer-supplied id from a trusted structured source
#             (e.g. IBM table [ROW_ID=HOST-1] → RFQ-HOST-001,
#              Nokia simple_list xlsx ID column "1" → RFQ-001)
#   - AI-NNN: system-generated when no trustworthy id exists
#             (e.g. HPE doc_schema rule="AI auto" → demote any LLM-hallucinated
#              RFQ-* to AI-NNN)
_AI_SEQ = {"n": 0}


def _reset_ai_seq():
    _AI_SEQ["n"] = 0


def _next_ai_id() -> str:
    """Return next AI-<NNN> with a single global counter."""
    _AI_SEQ["n"] += 1
    return f"AI-{_AI_SEQ['n']:03d}"


# ── Phase 4: doc_schema-driven req_id grounding ──────────────────────────────
# Goal: do NOT trust a "RFQ-HOST-NNN"-style req_id unless either:
#   (a) the source file's `req_id_rule` in inbound/<case>/meta/doc_schema.json
#       is a structured rule (e.g. "HOST N -> RFQ-HOST-{N:03d}, ..."), or
#   (b) the requirement text/notes carries a [ROW_ID=...] evidence marker that
#       confirms the prefix.
# Otherwise the id is treated as LLM-hallucinated and demoted to AI-<NNN>.

def load_doc_schema(enriched_in_path: Path) -> Dict[str, Any]:
    """Find inbound/<case>/meta/doc_schema.json given the enriched.json path.
    Returns {} if not found or unreadable. Best-effort path walk so callers can
    pass either an absolute or a working-dir-relative path to enriched.json."""
    p = Path(enriched_in_path)
    if not p.exists():
        return {}
    case_id = p.parent.name
    for ancestor in p.parents:
        candidate = ancestor / "inbound" / case_id / "meta" / "doc_schema.json"
        if candidate.exists():
            try:
                data = json.loads(candidate.read_text(encoding="utf-8"))
                return data if isinstance(data, dict) else {}
            except Exception:
                return {}
    return {}


def _file_req_id_rule(doc_schema: Dict[str, Any], filename: str) -> str:
    """Return the req_id_rule string for the named file.

    Prefers per-file rule from doc_schema.files[]; falls back to the legacy
    global rule. Returns "" when nothing is set.
    """
    if not doc_schema or not filename:
        return str((doc_schema or {}).get("req_id_rule") or "").strip()
    for f in (doc_schema.get("files") or []):
        if f.get("file") == filename:
            return str(f.get("req_id_rule") or "").strip()
    return str(doc_schema.get("req_id_rule") or "").strip()


def _is_structured_req_id_rule(rule: str) -> bool:
    """True iff the rule indicates the source file actually carries
    structured row IDs (not the placeholder 'AI auto')."""
    if not rule:
        return False
    norm = rule.strip().lower()
    return norm not in ("ai auto", "ai 自動編號", "ai 自动编号")


_ROW_ID_EVIDENCE_RE = re.compile(r"\[ROW_ID\s*=\s*([A-Za-z0-9_\-]+)\]", re.IGNORECASE)


def _has_row_id_evidence(text: str, notes: str, expected_prefix: str = "") -> bool:
    """True iff text or notes contains a [ROW_ID=PREFIX-N] marker
    (optionally narrowing to a specific prefix). The marker is the extractor's
    own breadcrumb left by read_docx_blocks(); its presence is reliable
    evidence that this id genuinely came from a structured table."""
    if not text and not notes:
        return False
    combined = f"{text or ''} {notes or ''}"
    for m in _ROW_ID_EVIDENCE_RE.finditer(combined):
        val = m.group(1)
        if not expected_prefix:
            return True
        if expected_prefix.upper() in val.upper():
            return True
    return False


def _is_placeholder(x: Any) -> bool:
    if x is None:
        return True
    if isinstance(x, str) and x.strip() in PLACEHOLDER_VALUES:
        return True
    return False


def _is_rfq_table_id(req_id: str) -> bool:
    for pat in _RFQ_TABLE_ID_PATTERNS:
        if pat.match(req_id.strip()):
            return True
    return False


def _safe_file_key(s: str) -> str:
    s = str(s).strip()
    s = re.sub(r"[\\/:*?\"<>|]+", "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:80] if len(s) > 80 else s


def guess_primary_source_file(reqs: List[Dict[str, Any]], meta: Dict[str, Any]) -> str:
    files = []
    for r in reqs:
        src = r.get("source", {}) or {}
        f = src.get("file")
        if _is_placeholder(f):
            continue
        f = str(f).strip()
        if re.fullmatch(r"\d+", f):
            continue
        files.append(f)

    if not files:
        return "UNKNOWN_SOURCE"

    docx = [f for f in files if f.lower().endswith(".docx")]
    if docx:
        return Counter(docx).most_common(1)[0][0]
    return Counter(files).most_common(1)[0][0]


def normalize_source(r: Dict[str, Any], fallback_file: str) -> Tuple[str, int]:
    src = r.get("source", {}) or {}
    f = src.get("file")
    c = src.get("chunk")

    if _is_placeholder(f):
        f = fallback_file
    else:
        f = str(f).strip()

    try:
        c_int = int(str(c).strip())
    except Exception:
        c_int = -1

    normalized: Dict[str, Any] = {"file": f, "chunk": c_int}
    # Preserve xlsx sheet / row if present
    if src.get("sheet"):
        normalized["sheet"] = str(src["sheet"]).strip()
    if src.get("row") is not None:
        try:
            normalized["row"] = int(src["row"])
        except (ValueError, TypeError):
            pass
    # Preserve docx table / table_row if present
    if src.get("table") is not None:
        try:
            normalized["table"] = int(src["table"])
        except (ValueError, TypeError):
            pass
    if src.get("table_row") is not None:
        try:
            normalized["table_row"] = int(src["table_row"])
        except (ValueError, TypeError):
            pass
    # Preserve pdf page if present
    if src.get("page") is not None:
        try:
            normalized["page"] = int(src["page"])
        except (ValueError, TypeError):
            pass
    r["source"] = normalized
    return f, c_int


def classify_item(text: str, notes: str) -> str:
    t = (text or "").strip()
    n = (notes or "").strip().upper()

    # 明確是 glossary
    if "GLOSSARY/DEFINITION" in n or "GLOSSARY" in n or "DEFINITION" in n:
        return "glossary"
    if re.match(r"^[A-Za-z0-9][A-Za-z0-9\-\+\s]{0,20}:\s+\S+", t):
        return "glossary"

    # 明確是 RFP 流程說明，不是需求
    if "RFP_PROCESS" in n:
        return "note"

    # 內容太短或空白
    word_count = len(re.findall(r"\w+", t))
    if not t:
        return "junk"
    if is_part_number_or_spec(t):
        return "junk"
    if word_count < 3 and not _REQUIREMENT_VERBS.search(t):
        return "junk"
    if word_count < 4:
        return "note"

    # 有明確需求動詞 → requirement
    if re.search(r"\b(must|shall|required|conform|comply|should|will|need|support|provide|ensure|enable|allow|implement|include|have|be capable|capable of)\b", t, re.IGNORECASE):
        return "requirement"

    # 8 字以上的句子，預設是 requirement（不是 note）
    if word_count >= 8:
        return "requirement"

    return "note"


def glue_orphan_replaceable(items: List[Dict[str, Any]]) -> None:
    prev_by_key = {}

    for r in items:
        src = r.get("source", {}) or {}
        key = (src.get("file"), src.get("chunk"))
        text = (r.get("requirement") or "").strip()

        if re.search(r"\bcan be replaced by\b", text, re.IGNORECASE):
            prev_text = (prev_by_key.get(key) or "").strip()
            m = re.match(r"^(CRU|FRU)\s*:\s*", prev_text, re.IGNORECASE)
            if m:
                prefix = m.group(1).upper()
                if not re.match(r"^(CRU|FRU)\b", text, re.IGNORECASE):
                    r["requirement"] = f"{prefix} {text}"
                    r["notes"] = (r.get("notes", "") + " | SERVICEABILITY_GLUE").strip(" |")

        prev_by_key[key] = text


_REQUIREMENT_VERBS = re.compile(
    r"\b(shall|must|required|obligation|deliverable)\b", re.IGNORECASE
)
_PART_NUMBER_ONLY = re.compile(
    r"^[A-Z0-9][-A-Z0-9\s\.]{0,25}$"
)
_BARE_NUMBER = re.compile(
    r"^[\d]+(\.\d+)?\s*(%|GB|MB|TB|GHz|MHz|V|W|A|mm|kg|lbs?|pcs?|ea)?$", re.IGNORECASE
)

# ── Phase 4: obvious-noise patterns ──────────────────────────────────────────
# Each pattern is anchored ^...$ so it only triggers when the WHOLE text is
# the noise — a real requirement that happens to contain a date/name as a
# substring will NOT be caught.

_PURE_DATE = re.compile(
    r"^\s*(?:"
    r"\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4}"            # 5/5/26, 5-5-2026, 5.5.26
    r"|\d{4}[/\-.]\d{1,2}[/\-.]\d{1,2}"             # 2026/05/05
    r"|\d{1,2}[\s\-/](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-/]\d{2,4}"
    r"|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},?\s+\d{2,4}"
    r")\s*$", re.IGNORECASE
)

_EMAIL_ONLY = re.compile(
    r"^\s*[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\s*$"
)

# "Carty, Clark <ck@x.com>" — name + email-in-angle-brackets
_CONTACT_WITH_EMAIL = re.compile(
    r"^\s*[A-Z][A-Za-z\s,.\-']*?<\s*[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\s*>\s*$"
)

_TICK_OR_LABEL = re.compile(
    r"^\s*(?:[✓✔✕✖✗✘×√]+"   # added √ (U+221A SQUARE ROOT) — common Asian-doc tick
    r"|Y|N|YES|NO|N/A|NA|TBD|TBC|TBA|N\.A\.|NIL|Y/N|N/Y)\s*$",
    re.IGNORECASE
)

# Single capitalized name, 1-2 title-case words, letters/hyphen only
# Will catch "Bernd", "Jing Hui". Won't catch ALL-CAPS like "CPU" or numeric "8GB".
_SINGLE_NAME = re.compile(
    r"^\s*[A-Z][a-z]+(?:[\s\-][A-Z][a-z]+)?\s*$"
)

# 1-4 letter ALL-CAPS abbrev alone (HW, BMC, IT, BIOS)
_UPPER_ABBREV = re.compile(r"^\s*[A-Z]{1,4}\s*$")

# Revision/update log phrase (only triggers when text is short overall, see is_obvious_noise)
_REVLOG_PATTERN = re.compile(
    r"^\s*(?:"
    r"(?:schedule|forecast|specification|spec|status|design|review|version|requirement)"
    r"(?:\s*&\s*\w+)?\s+"
    r"(?:updated|modified|approved|reviewed|locked|finalized|completed|added|removed|frozen)"
    r"|(?:updated|modified|reviewed|approved|created)\s+by\s+"
    r"|rev(?:ision)?\s*[:#]?\s*\d"
    r"|owner\s*[:#]\s*[A-Z]"
    r")", re.IGNORECASE
)

# Country + cert label, no verb (e.g. "Argentina S-Mark")
_COUNTRY_CERT = re.compile(
    r"^\s*(?:Argentina|Brazil|Mexico|China|India|Korea|Japan|Taiwan|USA|EU|UK|"
    r"Germany|France|Australia|Canada|Russia|Israel|Singapore|Indonesia|Vietnam|Thailand)\s+"
    r"(?:S-Mark|INMETRO|CCC|BIS|KC|VCCI|BSMI|FCC|CE|UL|RoHS|REACH|RCM|MIC|NCC)\s*$",
    re.IGNORECASE
)

# Sub-item bullet marker at start (a/b/c, 1./2., i./ii., -, •)
# The outer `\s+` handles the gap between marker and content, so the bullet
# alternative itself must NOT also consume whitespace.
_ORPHAN_MARKER = re.compile(
    r"^\s*(?:"
    r"\(?[a-z]\)"               # a), (a), (A)
    r"|\d+[\.\)]"               # 1., 1)
    r"|[ivxIVX]{1,4}\."         # ii., I.
    r"|[\-•*◦●·]"               # -, •, ◦, ●, ·
    r"|[a-z]\."                 # a., A.
    r")\s+", re.IGNORECASE
)


def is_obvious_noise(text: str, notes: str = "") -> bool:
    """Return True for text that is clearly NOT a requirement.

    Conservative — every pattern is anchored to the WHOLE text, so a real
    requirement that contains a date/name as a substring is preserved.

    Items already classified as glossary/note via `notes` are exempted so
    classify_item can route them to the right sheet.
    """
    n_upper = (notes or "").upper()
    if "GLOSSARY" in n_upper or "DEFINITION" in n_upper or "RFP_PROCESS" in n_upper:
        return False

    t = (text or "").strip()
    if not t:
        return True

    # 1. Pure date
    if _PURE_DATE.match(t):
        return True
    # 2. Email-only or contact-with-email
    if _EMAIL_ONLY.match(t) or _CONTACT_WITH_EMAIL.match(t):
        return True
    # 3. Tick / short label value (✓, Y/N, TBD, …)
    if _TICK_OR_LABEL.match(t):
        return True
    # 4. ALL-CAPS abbrev alone (HW, BMC, …)
    if _UPPER_ABBREV.match(t):
        return True
    # 5. Country + cert label only
    if _COUNTRY_CERT.match(t):
        return True
    # 6. Revision / update log (only if text is short)
    if len(re.findall(r"\w+", t)) <= 7 and _REVLOG_PATTERN.match(t):
        return True
    # 7. Single capitalized name (1-2 title-case words, no requirement verb)
    if _SINGLE_NAME.match(t) and not _REQUIREMENT_VERBS.search(t):
        return True

    return False


def is_likely_orphan_subitem(text: str, item_type: str, is_derived: bool) -> bool:
    """Narrow detection: text starts with a/b/c/1./2./bullet marker AND
    is short AND has no requirement verb.

    Per Phase 4 spec: ONLY a/b/c-style sub-items missing parent context.
    Not used for general junk — those are handled by is_obvious_noise.
    """
    if is_derived:
        return False
    if item_type != "requirement":
        return False
    t = (text or "").strip()
    if not t:
        return False
    if not _ORPHAN_MARKER.match(t):
        return False
    word_count = len(re.findall(r"\w+", t))
    # Numbered real requirement ("1. Memory shall be 8GB") has a verb → not orphan
    if _REQUIREMENT_VERBS.search(t):
        return False
    # Full sentences (>8 words) without verb are also unlikely to be orphans
    if word_count > 8:
        return False
    return True


def is_junk_short(text: str) -> bool:
    """True if text is too short to be a real requirement and has no requirement verbs."""
    t = (text or "").strip()
    word_count = len(re.findall(r"\w+", t))
    if word_count >= 3:
        return False
    if _REQUIREMENT_VERBS.search(t):
        return False
    return True


def is_part_number_or_spec(text: str) -> bool:
    """True if text looks like a standalone part number, version, or bare spec value."""
    t = (text or "").strip()
    if not t:
        return False
    if _BARE_NUMBER.match(t):
        return True
    if _PART_NUMBER_ONLY.match(t) and not any(c.islower() for c in t):
        return True
    return False


def dedup_requirements(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Remove exact duplicates by normalized requirement text. Keep first occurrence."""
    seen: dict = {}
    out: List[Dict[str, Any]] = []
    removed = 0
    for r in items:
        key = (r.get("requirement") or "").strip().lower()
        if not key:
            out.append(r)
            continue
        if key in seen:
            removed += 1
            continue
        seen[key] = True
        out.append(r)
    if removed:
        print(f"[POSTPROCESS] Dedup: removed {removed} exact duplicates ({len(out)} remaining)")
    return out


def filter_junk(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Remove items that are clearly not requirements:
    - obvious noise (dates / contacts / names / symbols / revision-log / cert label / abbrev)
    - junk-short text (< 3 words with no requirement verb)
    - standalone part numbers / spec values

    AUTO_SKIP-classified items (glossary by enricher) are exempted from the
    new obvious-noise filter so the glossary sheet stays intact.
    """
    out: List[Dict[str, Any]] = []
    removed_short = 0
    removed_pn = 0
    removed_noise = 0
    for r in items:
        if r.get("derived_requirement"):
            out.append(r)
            continue
        text = (r.get("requirement") or "").strip()
        notes = (r.get("notes") or "").strip()
        status_u = str(r.get("status") or "").upper()

        if status_u != "AUTO_SKIP" and is_obvious_noise(text, notes):
            removed_noise += 1
            continue
        if is_junk_short(text):
            removed_short += 1
            continue
        if is_part_number_or_spec(text):
            removed_pn += 1
            continue
        out.append(r)
    if removed_short or removed_pn or removed_noise:
        print(
            f"[POSTPROCESS] Filter: removed {removed_noise} obvious-noise, "
            f"{removed_short} too-short, {removed_pn} part-number/spec "
            f"({len(out)} remaining)"
        )
    return out


# ── Normalization layer ─────────────────────────────────────────────────────

# Prompt-trace patterns to strip from evidence/next_action/risk notes
_PROMPT_TRACES = re.compile(
    r"(falls under RF_\w+|per redflag|which indicates|implies potential risk"
    r"|Evidence required \(per redflag\)\.\s*)",
    re.IGNORECASE,
)

# Short risk tag from RF_* id or long message → canonical short tag
_RISK_TAG_MAP = {
    # RF_* IDs
    "RF_CERT":        "CERT",
    "RF_RELIABILITY": "RELIABILITY",
    "RF_PRICE":       "PRICING",
    "RF_SCHEDULE":    "SCHEDULE",
    "RF_IP":          "IP/LEGAL",
    "RF_ROADMAP":     "LIFECYCLE",
    "RF_GLOSSARY":    "GLOSSARY",
    # heuristic tags
    "SECURITY_TPM_COMPLIANCE": "CERT",
    "SUBJECTIVE_ACCEPTANCE":   "ACCEPTANCE",
    "FIELD_SERVICE_IMPACT":    "SERVICEABILITY",
    # Phase 4 orphan-subitem tag (passes through unchanged)
    "ORPHAN_SUBITEM":          "ORPHAN_SUBITEM",
}

# Map long Chinese redflag messages (from redflags.yaml) to short tags
_MSG_PREFIX_TO_TAG = [
    ("\u8a8d\u8b49/\u6cd5\u898f",         "CERT"),         # 認證/法規
    ("\u53ef\u9760\u5ea6/\u74b0\u5883",     "RELIABILITY"),  # 可靠度/環境
    ("\u50f9\u683c/\u4ed8\u6b3e",           "PRICING"),      # 價格/付款
    ("\u4ea4\u671f/\u7f70\u5247",           "SCHEDULE"),     # 交期/罰則
    ("\u6388\u6b0a/IP",                     "IP/LEGAL"),     # 授權/IP
    ("\u4f9b\u8ca8\u5e74\u9650",           "LIFECYCLE"),     # 供貨年限
    ("\u540d\u8a5e\u89e3\u91cb",           "GLOSSARY"),      # 名詞解釋
]

# Short English risk note per tag
_RISK_NOTE_MAP = {
    "CERT":           "Requires certification evidence or compliance declaration",
    "RELIABILITY":    "Requires test plan or reliability report",
    "PRICING":        "Pricing/payment terms — PM/Sales review required",
    "SCHEDULE":       "Lead time/penalty clause — PM review required",
    "IP/LEGAL":       "IP/licensing terms — Legal review required",
    "LIFECYCLE":      "Lifecycle/EOL clause — PM/supply chain review required",
    "GLOSSARY":       "Definition/abbreviation — informational only",
    "ACCEPTANCE":     "Subjective acceptance criteria — clarify with customer",
    "SERVICEABILITY": "Field service impact — confirm replacement policy",
    "ORPHAN_SUBITEM": "Possible orphan sub-item; verify parent context in original document",
}

# Category correction: keyword → canonical category
_CATEGORY_CORRECTIONS = {
    "Compliance": [
        "FCC", "EMC", "IEC", "IEEE1613", "PTCRB", "UL", "CE", "RoHS", "REACH",
        "EN55032", "EN55035", "CISPR", "IEC62368", "IEC60950", "NEBS", "GR-63",
        "safety", "regulatory", "homologation", "certification",
    ],
    "Commercial": [
        "warranty", "quotation", "EOL", "exclusivity", "launch schedule",
        "NDA", "indemnity", "liability", "penalty", "payment", "Incoterms",
        "pricing", "discount", "delivery",
    ],
    "Reliability": [
        "shock", "vibration", "burn-in", "seismic", "HALT", "ALT", "MTBF",
        "humidity", "thermal cycling", "altitude", "acoustic",
    ],
}

# Expanded status vocabulary
VALID_STATUSES = {
    "NEW", "NEED_REVIEW", "INTERNAL_ALIGN", "ASK_CUSTOMER",
    "READY_FOR_RESPONSE", "CLOSED", "PENDING", "AUTO_SKIP",
}


def clean_llm_text(text: str) -> str:
    """Remove internal prompt traces from LLM-generated text."""
    if not text:
        return ""
    t = _PROMPT_TRACES.sub("", text).strip()
    t = re.sub(r"\s{2,}", " ", t)
    t = t.strip(" .,;")
    return t


_RF_ID_PATTERN = re.compile(r"RF_(CERT|RELIABILITY|PRICE|SCHEDULE|IP|ROADMAP|GLOSSARY)", re.IGNORECASE)

_KEYWORD_TO_TAG = [
    (r"\b(certif\w*|regulatory|FIPS|IEC6\w*|GB4943|UL\b|CE mark|FCC\b|RoHS|REACH|homologat\w*|safety report|EN55\w*|CISPR|NEBS|PTCRB)", "CERT"),
    (r"\b(reliab\w*|MTBF|HALT\b|ALT\b|vibrat\w*|shock|humid\w*|thermal.cycl\w*|burn.in|seismic|altitude)", "RELIABILITY"),
    (r"\b(price|pricing|cost\b|USD|TWD|payment|discount|Incoterms|quotat\w*)", "PRICING"),
    (r"\b(lead.time|deliver\w*|shipment|deadline|penalty|liquidated|SLA\b)", "SCHEDULE"),
    (r"\b(intellectual property|IP rights|licens\w*|escrow|source code|redistribut\w*)", "IP/LEGAL"),
    (r"\b(roadmap|lifecycle|EOL\b|end of life|7.year|availab\w*)", "LIFECYCLE"),
]


def _resolve_tag(text: str) -> Optional[str]:
    """Try to resolve a raw tag/message string to a canonical short tag."""
    t = text.strip()
    if not t:
        return None
    if t in _RISK_TAG_MAP:
        return _RISK_TAG_MAP[t]
    for prefix, short in _MSG_PREFIX_TO_TAG:
        if t.startswith(prefix):
            return short
    m = _RF_ID_PATTERN.search(t)
    if m:
        rf_id = f"RF_{m.group(1).upper()}"
        return _RISK_TAG_MAP.get(rf_id)
    for pattern, short in _KEYWORD_TO_TAG:
        if re.search(pattern, t, re.IGNORECASE):
            return short
    return None


def normalize_risk_flags(raw_tags: list) -> Tuple[List[str], str]:
    """Convert raw RF_*/long-message tags into (short_tags, one_line_note)."""
    short_tags = []
    seen = set()
    for tag in raw_tags:
        canonical = _resolve_tag(str(tag))
        if canonical:
            if canonical not in seen:
                short_tags.append(canonical)
                seen.add(canonical)
        else:
            t = str(tag).strip()
            if t and t not in seen:
                short_tags.append(t)
                seen.add(t)

    if short_tags:
        notes = [_RISK_NOTE_MAP.get(t, "") for t in short_tags]
        note = "; ".join(n for n in notes if n)
    else:
        note = ""
    return short_tags, note


def normalize_category(category_str: str, requirement_text: str) -> str:
    """Rule-based correction: override General with specific category if keywords match."""
    if category_str and category_str != "General":
        return category_str
    tl = (requirement_text or "").lower()
    for cat, keywords in _CATEGORY_CORRECTIONS.items():
        for kw in keywords:
            if kw.lower() in tl:
                return cat
    return category_str or "General"


def normalize_status(raw_status: str, item_type: str) -> str:
    """Map old status values to extended vocabulary."""
    s = (raw_status or "").strip().upper()
    if item_type == "glossary":
        return "AUTO_SKIP"
    if item_type in ("note", "junk"):
        return "AUTO_SKIP"
    if s in ("PENDING", "AUTO_OK", ""):
        return "NEW"
    if s in VALID_STATUSES:
        return s
    return "NEW"


def split_pipe_requirement(req_text: str) -> List[str]:
    parts = [x.strip() for x in (req_text or "").split("|")]
    parts = [p for p in parts if p]
    if len(parts) <= 1:
        return [(req_text or "").strip()]
    return parts


def must_level_heuristic(text: str, item_type: str) -> str:
    if item_type == "glossary":
        return "INFO"
    t = (text or "").lower()
    if "must" in t:
        return "MUST"
    if "shall" in t:
        return "MUST"
    if "can be replaced" in t:
        return "MAY"
    return "INFO"


def auto_owner_category_redflags_heuristic(text: str, item_type: str) -> Tuple[str, str, List[str], str, str]:
    if item_type != "requirement":
        return ("General", "TBD", [], "", "")

    tl = (text or "").lower()
    redflags = []
    evidence = ""
    next_action = ""

    if re.search(r"\b(tpm|trusted computing|tcg)\b", tl):
        return ("Security", "BIOS/Security", ["SECURITY_TPM_COMPLIANCE"],
                "TPM implementation (dTPM/fTPM), part datasheet, certification scope",
                "Confirm TPM implementation and compliance scope")

    if re.search(r"\b(bezel|lcd|keypad|button|led|front bezel|rear bezel|console port)\b", tl):
        category = "Mechanical"
        owner = "ME/ID"
        if re.search(r"\b(look and feel|common)\b", tl) or "required" in tl:
            redflags.append("SUBJECTIVE_ACCEPTANCE")
            evidence = "Customer acceptance criteria / reference product / mechanical spec"
            next_action = "Clarify acceptance criteria with customer"
        else:
            evidence = "Mechanical spec / ID drawing / BOM impact"
            next_action = "Assess feasibility & cost impact"
        return (category, owner, redflags, evidence, next_action)

    if re.search(r"\b(fru|cru|replaceable|can be replaced)\b", tl):
        return ("Serviceability", "QA/Service", ["FIELD_SERVICE_IMPACT"],
                "Service policy / replacement procedure / responsibility boundary",
                "Confirm replacement policy (customer vs certified technician)")

    return ("General", "TBD", redflags, evidence, next_action)


def auto_status(item_type: str, mlevel: str, redflags: List[str]) -> str:
    if item_type == "glossary":
        return "AUTO_SKIP"
    if item_type != "requirement":
        return "AUTO_SKIP"
    if mlevel in ("MUST", "SHALL"):
        return "NEED_REVIEW"
    if redflags:
        return "NEED_REVIEW"
    if mlevel == "MAY":
        return "NEED_REVIEW"
    return "NEW"


def ensure_req_id(
    r: Dict[str, Any],
    file_name: str,
    chunk: int,
    doc_schema: Optional[Dict[str, Any]] = None,
) -> Tuple[str, str, str]:
    """Resolve the final req_id with Phase 4 grounding validation.

    Returns (req_id, id_source, original_id) where id_source is:
      - "original"  : id is backed by a trusted structured source
                      (doc_schema.req_id_rule is structured) OR by a
                      [ROW_ID=...] evidence marker in the chunk text.
      - "generated" : no trustworthy id; fresh AI-<NNN>.

    Output formats:
      - RFQ-<PREFIX>-<NNN>  : trusted PREFIX-N table-row id  (IBM, ...)
      - RFQ-<NNN>           : trusted numeric raw id from a structured source
                              (Nokia simple_list xlsx ID column)
      - RFQ-TBL-<NNN>       : trusted ROW_ID=N marker
      - AI-<NNN>            : system-generated; covers LLM-hallucinated RFQ-*
                              (HPE), unknown raw ids, and AUTO-* placeholders.

    Special inputs:
      __id_locked__=True  — outer split-pipe call already resolved this id;
                            preserve to avoid re-validating slice text.
    """
    # Pre-locked from outer split-pipe recursion: short-circuit
    if r.get("__id_locked__"):
        return (
            str(r.get("req_id", "")),
            str(r.get("__id_source__", "generated")),
            str(r.get("__id_original__", r.get("req_id", ""))),
        )

    if doc_schema is None:
        doc_schema = {}

    rid = r.get("req_id")
    rid_str = str(rid).strip() if rid is not None else ""
    original_id = rid_str
    text = r.get("requirement", "")
    notes = r.get("notes", "")

    file_rule = _file_req_id_rule(doc_schema, file_name)
    is_structured = _is_structured_req_id_rule(file_rule)

    # Case 1: raw RFQ table id  (HOST-30 / BMC-1 / ROW_ID=2)
    if rid_str and _is_rfq_table_id(rid_str):
        m = re.match(r"^([A-Za-z_]+)-(\d+)$", rid_str)
        if m:
            prefix, num = m.group(1), m.group(2)
            if is_structured or _has_row_id_evidence(text, notes, prefix):
                return (f"RFQ-{prefix.upper()}-{int(num):03d}", "original", original_id)
            return (_next_ai_id(), "generated", original_id)
        m2 = re.match(r"^ROW_ID=(\d+)$", rid_str)
        if m2:
            if is_structured or _has_row_id_evidence(text, notes, ""):
                return (f"RFQ-TBL-{int(m2.group(1)):03d}", "original", original_id)
            return (_next_ai_id(), "generated", original_id)
        if is_structured:
            return (f"RFQ-{rid_str}", "original", original_id)
        return (_next_ai_id(), "generated", original_id)

    # Case 2a: pre-formatted RFQ-PREFIX-NNN  (letter-prefix form, with optional suffix)
    if rid_str.startswith("RFQ-"):
        m_letter = re.match(r"^RFQ-([A-Za-z_]+)-\d+(?:-[a-z0-9]+)?$", rid_str)
        if m_letter:
            prefix = m_letter.group(1)
            if is_structured or _has_row_id_evidence(text, notes, prefix):
                return (rid_str, "original", original_id)
            return (_next_ai_id(), "generated", original_id)
        # Numeric form RFQ-NNN  (with optional suffix) — only valid if structured
        m_num = re.match(r"^RFQ-\d+(?:-[a-z0-9]+)?$", rid_str)
        if m_num:
            if is_structured:
                return (rid_str, "original", original_id)
            return (_next_ai_id(), "generated", original_id)
        # Unparseable RFQ-* — preserve cautiously
        return (rid_str, "original", original_id)

    # Case 2b: existing AI-NNN[-suffix]  →  keep as generated
    if re.match(r"^AI-\d+(?:-[a-z0-9]+)?$", rid_str):
        return (rid_str, "generated", original_id)

    # Case 2.5: structured numeric raw id  (Nokia simple_list xlsx ID column: "1" → RFQ-001)
    if rid_str and is_structured and rid_str.isdigit():
        return (f"RFQ-{int(rid_str):03d}", "original", original_id)

    # Case 3: AUTO-* / empty / unknown / unstructured-non-PREFIX → fresh AI-<NNN>
    return (_next_ai_id(), "generated", original_id)


def is_enriched_item(r: Dict[str, Any]) -> bool:
    return any(k in r for k in ("must_level", "owner", "status", "redflag_messages", "redflag_overrides", "category"))


def flatten_redflags(r: Dict[str, Any]) -> List[str]:
    tags = []
    msgs = r.get("redflag_messages") or []
    if isinstance(msgs, list):
        for m in msgs:
            m = str(m).strip()
            if m:
                tags.append(m)
    rt = r.get("redflag_tags") or []
    if isinstance(rt, list):
        tags.extend([str(x).strip() for x in rt if str(x).strip()])
    return tags


def build_rows(items: List[Dict[str, Any]], doc_schema: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if doc_schema is None:
        doc_schema = {}

    for r in items:
        src = r.get("source", {}) or {}
        f = src.get("file", "UNKNOWN_SOURCE")
        c = src.get("chunk", -1)

        text = (r.get("requirement") or "").strip()
        notes = (r.get("notes") or "").strip()
        conf = r.get("confidence", "")

        # Split pipe rows
        parts = split_pipe_requirement(text)
        if len(parts) > 1:
            # Resolve once for the parent row; recursion sees __id_locked__
            # and skips re-validation against the (smaller) slice text.
            base_id, base_id_src, base_orig = ensure_req_id(
                r, f, c, doc_schema=doc_schema
            )
            for idx, p in enumerate(parts):
                rr = dict(r)
                rr["requirement"] = p
                suffix = chr(ord("a") + idx) if idx < 26 else str(idx + 1)
                rr["req_id"] = f"{base_id}-{suffix}"
                rr["__id_locked__"]   = True
                rr["__id_source__"]   = base_id_src
                rr["__id_original__"] = (f"{base_orig}-{suffix}" if base_orig else rr["req_id"])
                rr["notes"] = (notes + " | SPLIT_FROM_PIPE").strip(" |")
                rows.extend(build_rows([rr], doc_schema=doc_schema))
            continue

        is_derived = bool(r.get("derived_requirement"))
        item_type = "requirement" if is_derived else classify_item(text, notes)

        # 強制覆蓋：LLM 已判斷為重要的需求，不管文字多短都是 requirement
        must_level_val = str(r.get("must_level") or "").upper()
        status_val = str(r.get("status") or "")
        redflag_val = r.get("redflag_tags") or r.get("redflag_messages") or []
        has_rf = bool(redflag_val and str(redflag_val) not in ("[]", ""))

        if item_type in ("note", "junk"):
            # Phase 4 guard: obvious-noise text is never promoted to requirement,
            # even if upstream LLM enrichment marked it MUST/SHOULD/MAY.
            if is_obvious_noise(text, notes):
                item_type = "junk"
            elif must_level_val in ("MUST", "SHOULD", "MAY"):
                item_type = "requirement"
            elif status_val == "NEED_REVIEW":
                item_type = "requirement"
            elif has_rf:
                item_type = "requirement"
            elif status_val == "AUTO_SKIP":
                item_type = "glossary"

        enriched = is_enriched_item(r)

        if is_derived:
            # Derived requirements from spec_reference: use spec_category, default NEED_REVIEW
            category = r.get("spec_category") or "General"
            mlevel = "INFO"
            owner = "TBD"
            status = "NEED_REVIEW"
            raw_rf = []
            evidence = ""
            next_action = "Confirm spec against design. Derived from spec table, not explicit requirement."
        elif enriched:
            mlevel = r.get("must_level") or must_level_heuristic(text, item_type)
            cat = r.get("category") or "General"
            category = ", ".join(cat) if isinstance(cat, list) else str(cat)
            owner = r.get("owner") or "TBD"
            status = r.get("status") or auto_status(item_type, str(mlevel), [])
            raw_rf = flatten_redflags(r)
            evidence = ""
            next_action = ""
            overrides = r.get("redflag_overrides") or {}
            if overrides.get("require_evidence") is True:
                evidence = "Provide datasheet/report/certification as applicable."
                next_action = "Collect evidence and confirm with owner."
        else:
            mlevel = must_level_heuristic(text, item_type)
            category, owner, redflags, evidence, next_action = auto_owner_category_redflags_heuristic(text, item_type)
            status = auto_status(item_type, mlevel, redflags)
            raw_rf = redflags

        # Phase 4: narrow orphan-subitem detection (a/b/c bullet without parent context)
        if is_likely_orphan_subitem(text, item_type, is_derived):
            if "ORPHAN_SUBITEM" not in raw_rf:
                raw_rf.append("ORPHAN_SUBITEM")
            status = "NEED_REVIEW"
            if not next_action:
                next_action = "Verify parent context — read source document"

        # ── Normalize ──
        category = normalize_category(category, text)
        risk_tags, risk_note = normalize_risk_flags(raw_rf)
        status = normalize_status(status, item_type)
        evidence = clean_llm_text(evidence)
        next_action = clean_llm_text(next_action)

        rid, id_source, original_id = ensure_req_id(
            r, f, c, doc_schema=doc_schema
        )
        sheet = src.get("sheet", "")
        row = src.get("row")
        fname = re.sub(r'\.(docx|doc|xlsx|xls|pdf|txt|md)$', '', Path(f).name, flags=re.IGNORECASE)
        fname = re.sub(r'^AUTO[-_]?', '', fname, flags=re.IGNORECASE).strip()
        fname = re.sub(r'^(UNKNOWN|UnknownFile)$', '', fname, flags=re.IGNORECASE).strip()
        fname = fname[:40]
        table = src.get("table")
        table_row = src.get("table_row")
        page = src.get("page")

        if not fname:
            source_ref = ""
        elif sheet and row is not None:
            source_ref = f"{fname} \u2014 Sheet: {sheet}, \u7b2c {row} \u884c"
        elif sheet:
            source_ref = f"{fname} \u2014 Sheet: {sheet}"
        elif table is not None and table_row is not None:
            source_ref = f"{fname} \u2014 Table {table}, Row {table_row}"
        elif page is not None:
            source_ref = f"{fname} \u2014 \u7b2c {page} \u9801"
        elif c is not None and c >= 0:
            source_ref = f"{fname} \u2014 \u7b2c {c} \u6bb5"
        else:
            source_ref = fname

        sh = r.get("stakeholder") or []
        stakeholder = ", ".join(sh) if isinstance(sh, list) else str(sh)
        rows.append({
            "Req ID": rid,
            "重要程度 (Must Level)": str(mlevel),
            "Category": category,
            "Owner": owner,
            "Stakeholder": stakeholder,
            "Status": status,
            "Requirement (Original)": text,
            "Risk Flags / 風險標記": ", ".join(risk_tags) if risk_tags else "",
            "Risk Note": risk_note,
            "Evidence Notes": evidence,
            "Next Action": next_action,
            "Source": source_ref,
            "_orig_req_id": r.get("req_id", ""),
            "_type": item_type,
            "_derived": is_derived,
            "_id_source": id_source,
            "_original_id": original_id,
        })

    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="in_path", required=True)
    ap.add_argument("--out_dir", default=".")
    ap.add_argument("--xlsx_name", default="requirements_review.xlsx")
    args = ap.parse_args()

    in_path = Path(args.in_path)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    data = json.loads(in_path.read_text(encoding="utf-8"))
    meta = data.get("meta", {}) or {}
    reqs = data.get("requirements") or data.get("items") or []

    fallback_file = guess_primary_source_file(reqs, meta)

    for r in reqs:
        normalize_source(r, fallback_file)

    glue_orphan_replaceable(reqs)

    print(f"[POSTPROCESS] Before cleaning: {len(reqs)} requirements")
    reqs = dedup_requirements(reqs)
    reqs = filter_junk(reqs)
    print(f"[POSTPROCESS] After cleaning: {len(reqs)} requirements")

    # Phase 4: load doc_schema so ensure_req_id can validate RFQ-* claims.
    # When missing/unreadable an empty dict is returned — every RFQ-* claim
    # then defaults to "untrusted" and is demoted to AI-<NNN>.
    doc_schema = load_doc_schema(in_path)
    _ds_rule = (doc_schema or {}).get("req_id_rule", "")
    _ds_files = len((doc_schema or {}).get("files") or [])
    print(f"[POSTPROCESS] doc_schema loaded: req_id_rule={_ds_rule!r} files={_ds_files}")

    _reset_ai_seq()
    rows = build_rows(reqs, doc_schema=doc_schema)

    cols = [
        "Req ID", "重要程度 (Must Level)", "Category", "Owner", "Stakeholder", "Status",
        "Requirement (Original)", "Risk Flags / 風險標記", "Risk Note",
        "Evidence Notes", "Next Action", "Source",
    ]
    _internal_cols = cols + ["_orig_req_id", "_type"]
    hidden_cols = {"Source"}

    if rows:
        df = pd.DataFrame(rows)
    else:
        df = pd.DataFrame(columns=_internal_cols)
        print("[POSTPROCESS] No requirements found after cleaning. "
              "Current files may be spec-reference / checklist documents without explicit requirements. "
              "Outputting empty templates.")

    df_req  = df[df["_type"] == "requirement"].copy() if "_type" in df.columns else pd.DataFrame(columns=_internal_cols)
    df_glo  = df[df["_type"] == "glossary"].copy()    if "_type" in df.columns else pd.DataFrame(columns=_internal_cols)
    df_note = df[df["_type"].isin(["note", "junk"])].copy() if "_type" in df.columns else pd.DataFrame(columns=_internal_cols)

    df_req  = df_req.reindex(columns=cols, fill_value="")
    df_glo  = df_glo.reindex(columns=cols, fill_value="")
    df_note = df_note.reindex(columns=cols, fill_value="")

    # ── PM 排序：NEED_REVIEW 先 > 有 Redflag 先 > MUST 先 > Category ──────────
    must_order   = {"MUST": 0, "SHOULD": 1, "MAY": 2, "INFO": 3}
    status_order = {"NEED_REVIEW": 0, "INTERNAL_ALIGN": 1, "ASK_CUSTOMER": 2,
                    "NEW": 3, "READY_FOR_RESPONSE": 4, "CLOSED": 5, "AUTO_SKIP": 6}
    cat_order    = {
        "Compliance": 0, "Reliability": 1, "Security": 2,
        "Platform": 3, "BMC": 4, "BIOS": 5, "Storage": 6,
        "PCIe": 7, "Power": 8, "Thermal": 9, "Mechanical": 10, "General": 99,
    }
    if not df_req.empty:
        df_req = df_req.copy()
        df_req["_sort_status"] = df_req["Status"].map(status_order).fillna(9)
        df_req["_sort_rf"]     = df_req["Risk Flags / 風險標記"].apply(lambda x: 0 if pd.notna(x) and str(x).strip() != "" else 1)
        df_req["_sort_must"]   = df_req["重要程度 (Must Level)"].map(must_order).fillna(9)
        df_req["_sort_cat"]    = df_req["Category"].map(cat_order).fillna(50)
        df_req = df_req.sort_values(["_sort_status","_sort_rf","_sort_must","_sort_cat"]).drop(
            columns=["_sort_status","_sort_rf","_sort_must","_sort_cat"]
        ).reset_index(drop=True)

    xlsx_path = out_dir / args.xlsx_name
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_req.to_excel(writer,  sheet_name="Requirements", index=False)
        df_glo.to_excel(writer,  sheet_name="Glossary",     index=False)
        df_note.to_excel(writer, sheet_name="Notes",        index=False)

        from openpyxl.styles import PatternFill, Font
        for sheet_name in ("Requirements", "Glossary", "Notes"):
            ws = writer.book[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{ws.cell(1, ws.max_column).column_letter}{ws.max_row}"

            # Hide Source column
            for col_idx, header in enumerate(cols, start=1):
                if header in hidden_cols:
                    ws.column_dimensions[ws.cell(1, col_idx).column_letter].hidden = True

            # Colour-code Req ID: RFQ- green, AI- blue
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=1)
                val = str(cell.value or "")
                if val.startswith("RFQ-"):
                    cell.fill = PatternFill("solid", fgColor="EAF3DE")
                    cell.font = Font(color="27500A", name="Arial", size=10)
                elif val.startswith("AI-"):
                    cell.fill = PatternFill("solid", fgColor="E6F1FB")
                    cell.font = Font(color="0C447C", name="Arial", size=10)

    # Write clean.json
    _summary = {
        "total_requirements": len(df_req),
        "total_glossary": len(df_glo),
        "total_notes": len(df_note),
    }
    if len(df_req) == 0 and len(df_glo) == 0:
        _summary["pm_note"] = (
            "No actionable requirements extracted. "
            "Uploaded files appear to be spec-reference, datasheet, or checklist documents "
            "without explicit shall/must requirements."
        )
    clean = {
        "meta": meta,
        "postprocess": {
            "input": str(in_path.name),
            "fallback_source_file": fallback_file,
            "xlsx": str(xlsx_path.name),
            "columns": cols,
        },
        "summary": _summary,
        "items": []
    }

    _export_cols = [c for c in cols + ["_type", "_orig_req_id", "_derived", "_id_source", "_original_id"] if c in df.columns]
    for r in df.reindex(columns=_export_cols, fill_value="").to_dict(orient="records"):
        _rf_raw = r.get("Risk Flags / 風險標記") or ""
        _rf_list = [t.strip() for t in _rf_raw.split(",") if t.strip()] if _rf_raw else []
        clean["items"].append({
            "req_id":        r["Req ID"],
            "orig_req_id":   r.get("_orig_req_id", ""),
            "original_id":   r.get("_original_id", ""),
            "id_source":     r.get("_id_source", "unknown"),
            "type":          r.get("_type", ""),
            "must_level":    r.get("重要程度 (Must Level)", ""),
            "category":      r.get("Category", ""),
            "owner":         r.get("Owner", ""),
            "stakeholder":   [s.strip() for s in r["Stakeholder"].split(",") if s.strip()] if r.get("Stakeholder") else [],
            "status":        r.get("Status", ""),
            "requirement":   r.get("Requirement (Original)", ""),
            "risk_tags":     _rf_list,
            "risk_note":     r.get("Risk Note", ""),
            "evidence_needed": r.get("Evidence Notes", ""),
            "next_action":   r.get("Next Action", ""),
            "source":        r.get("Source", ""),
            "derived":       bool(r.get("_derived", False)),
            # ── Phase 4.6 normalization fields (filled by
            # scripts/normalize_requirements_llm.py; empty defaults keep the
            # clean.json schema consistent across all cases). ──
            "normalized_requirement": "",
            "rewrite_reason":         "",
            "rewrite_confidence":     0.0,
            "needs_rewrite_review":   False,
        })

    clean_path = out_dir / "requirements_clean.json"
    clean_path.write_text(json.dumps(clean, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"[OK] Wrote: {xlsx_path}")
    print(f"[OK] Wrote: {clean_path}")
    print(f"[INFO] fallback_source_file: {fallback_file}")
    print(f"[INFO] counts: Requirements={len(df_req)}, Glossary={len(df_glo)}, Notes={len(df_note)}")


if __name__ == "__main__":
    main()
