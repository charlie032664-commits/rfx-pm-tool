# -*- coding: utf-8 -*-
"""Compare two pipeline run outputs (v1.3) — statistics only, no LLM.

Compares EXISTING outputs of two run directories (e.g. the same case extracted
by two different models/providers) and writes a markdown report. It reports
counts and distributions only — it never dumps requirement / customer text.

Usage:
    python compare_model_outputs.py --left runs/<caseA> --right runs/<caseB> \
        --out runs/_compare_report.md
"""
from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List

STD_FILES = [
    "manifest.json", "requirements.json", "requirements_enriched.json",
    "requirements_clean.json", "requirements_review.xlsx", "compliance_matrix.xlsx",
]


def _load_json(p: Path):
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None


def _cat_of(item: Dict[str, Any]) -> str:
    c = item.get("category")
    if isinstance(c, list):
        c = c[0] if c else ""
    return str(c).strip() or "(none)"


def _matrix_rows(xlsx: Path) -> Any:
    if not xlsx.exists():
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
        sheet = "Compliance Matrix" if "Compliance Matrix" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        rows = max(ws.max_row - 1, 0)  # minus header
        wb.close()
        return rows
    except Exception:
        return None


def collect_stats(run_dir: Path) -> Dict[str, Any]:
    run_dir = Path(run_dir)
    d: Dict[str, Any] = {"run_dir": str(run_dir)}
    d["files"] = {f: (run_dir / f).exists() for f in STD_FILES}

    clean = _load_json(run_dir / "requirements_clean.json") or {}
    meta = clean.get("meta", {}) or {}
    d["model"] = str(meta.get("model", "") or "")
    items: List[Dict[str, Any]] = clean.get("items", []) or []
    reqs = [i for i in items if i.get("type") == "requirement"]

    d["total_items"] = len(items)
    d["requirements"] = len(reqs)
    d["glossary"] = len([i for i in items if i.get("type") == "glossary"])
    d["notes"] = len([i for i in items if i.get("type") in ("note", "junk")])

    ids = [str(i.get("req_id", "") or "") for i in items]
    d["duplicate_ids"] = sorted({x for x in ids if x and ids.count(x) > 1})
    d["empty_requirement_text"] = sum(1 for i in reqs if not str(i.get("requirement", "") or "").strip())
    d["empty_category"] = sum(1 for i in reqs if _cat_of(i) == "(none)")
    d["empty_owner"] = sum(1 for i in reqs if not str(i.get("owner", "") or "").strip())

    d["category_dist"] = dict(Counter(_cat_of(i) for i in reqs))
    d["must_level_dist"] = dict(Counter(str(i.get("must_level", "") or "(none)") for i in reqs))
    d["status_dist"] = dict(Counter(str(i.get("status", "") or "(none)") for i in reqs))

    # field completeness: expected keys present on every requirement item
    expected = ["req_id", "requirement", "category", "owner", "must_level", "status"]
    missing = {k: sum(1 for i in reqs if k not in i) for k in expected}
    d["missing_fields"] = {k: v for k, v in missing.items() if v}

    d["matrix_rows"] = _matrix_rows(run_dir / "compliance_matrix.xlsx")
    return d


def _dist_table(name: str, left: Dict[str, int], right: Dict[str, int]) -> str:
    keys = sorted(set(left) | set(right))
    lines = [f"### {name}", "", "| key | left | right |", "|---|---|---|"]
    for k in keys:
        lines.append(f"| {k} | {left.get(k, 0)} | {right.get(k, 0)} |")
    return "\n".join(lines)


def build_report(left: Dict[str, Any], right: Dict[str, Any]) -> str:
    def row(label, key):
        return f"| {label} | {left.get(key)} | {right.get(key)} |"

    out = []
    out.append("# Model Output Comparison (statistics only)\n")
    out.append(f"- **left:** `{left['run_dir']}` · model `{left.get('model') or '?'}`")
    out.append(f"- **right:** `{right['run_dir']}` · model `{right.get('model') or '?'}`\n")

    out.append("## Counts\n")
    out.append("| metric | left | right |")
    out.append("|---|---|---|")
    for label, key in [
        ("total items", "total_items"), ("requirements", "requirements"),
        ("glossary", "glossary"), ("notes", "notes"),
        ("compliance matrix rows", "matrix_rows"),
        ("empty requirement text", "empty_requirement_text"),
        ("empty category", "empty_category"), ("empty owner", "empty_owner"),
    ]:
        out.append(row(label, key))
    out.append("")

    out.append("## Duplicate req_ids\n")
    out.append(f"- left: {len(left['duplicate_ids'])} → {left['duplicate_ids'][:10]}")
    out.append(f"- right: {len(right['duplicate_ids'])} → {right['duplicate_ids'][:10]}\n")

    out.append("## Missing fields (count of requirement items lacking a key)\n")
    out.append(f"- left: {left['missing_fields'] or 'none'}")
    out.append(f"- right: {right['missing_fields'] or 'none'}\n")

    out.append(_dist_table("Category distribution", left["category_dist"], right["category_dist"]) + "\n")
    out.append(_dist_table("Must-level distribution", left["must_level_dist"], right["must_level_dist"]) + "\n")
    out.append(_dist_table("Status distribution", left["status_dist"], right["status_dist"]) + "\n")

    out.append("## Output file availability\n")
    out.append("| file | left | right |")
    out.append("|---|---|---|")
    for f in STD_FILES:
        out.append(f"| {f} | {'✅' if left['files'].get(f) else '—'} | {'✅' if right['files'].get(f) else '—'} |")
    out.append("")
    return "\n".join(out)


def main() -> int:
    ap = argparse.ArgumentParser(description="Compare two run outputs (no LLM).")
    ap.add_argument("--left", required=True, help="Left run dir")
    ap.add_argument("--right", required=True, help="Right run dir")
    ap.add_argument("--out", required=True, help="Markdown report path")
    args = ap.parse_args()

    left = collect_stats(Path(args.left))
    right = collect_stats(Path(args.right))
    report = build_report(left, right)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(report, encoding="utf-8")
    print(f"[OK] wrote comparison report: {out_path}")
    print(f"[INFO] left requirements={left['requirements']} right requirements={right['requirements']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
